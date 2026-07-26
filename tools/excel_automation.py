import openpyxl
import os
import json
from datetime import datetime
from typing import List, Dict

class ExcelAutomator:
    def __init__(self, input_file: str):
        self.input_file = input_file
        self.wb = openpyxl.load_workbook(input_file)
    
    def extract_data(self, sheet_name: str = None) -> List[Dict]:
        if sheet_name:
            sheet = self.wb[sheet_name]
        else:
            sheet = self.wb.active
        
        headers = [cell.value for cell in sheet[1]]
        data = []
        
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for idx, cell in enumerate(row):
                row_data[headers[idx]] = cell.value
            data.append(row_data)
        
        return data
    
    def generate_report(self, data: List[Dict], output_file: str, summary_columns: List[str] = None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        if data:
            headers = list(data[0].keys())
            ws.append(headers)
            
            for row in data:
                ws.append([row.get(h) for h in headers])
            
            if summary_columns:
                summary_ws = wb.create_sheet("Summary")
                summary_ws.append(["Column", "Count", "Unique Values", "Sum", "Avg"])
                
                for col in summary_columns:
                    values = [row[col] for row in data if row[col] is not None and isinstance(row[col], (int, float))]
                    unique = len(set([row[col] for row in data if row[col] is not None]))
                    
                    if values:
                        summary_ws.append([col, len(values), unique, sum(values), sum(values)/len(values)])
                    else:
                        summary_ws.append([col, len([row[col] for row in data if row[col] is not None]), unique, 0, 0])
        
        wb.save(output_file)
        print(f"Report saved to {output_file}")
    
    def batch_process_files(self, input_dir: str, output_dir: str):
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        all_data = []
        
        for filename in os.listdir(input_dir):
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                filepath = os.path.join(input_dir, filename)
                try:
                    temp_wb = openpyxl.load_workbook(filepath)
                    sheet = temp_wb.active
                    headers = [cell.value for cell in sheet[1]]
                    
                    for row in sheet.iter_rows(min_row=2):
                        row_data = {}
                        for idx, cell in enumerate(row):
                            row_data[headers[idx]] = cell.value
                        row_data['source_file'] = filename
                        all_data.append(row_data)
                    
                    print(f"Processed: {filename}")
                except Exception as e:
                    print(f"Error processing {filename}: {e}")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(output_dir, f"combined_report_{timestamp}.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Combined Data"
        
        if all_data:
            headers = list(all_data[0].keys())
            ws.append(headers)
            
            for row in all_data:
                ws.append([row.get(h) for h in headers])
        
        wb.save(output_file)
        print(f"Combined report saved to {output_file}")
        return len(all_data)
    
    def find_duplicates(self, column_name: str, sheet_name: str = None) -> List[Dict]:
        if sheet_name:
            sheet = self.wb[sheet_name]
        else:
            sheet = self.wb.active
        
        headers = [cell.value for cell in sheet[1]]
        col_idx = headers.index(column_name)
        
        values = {}
        duplicates = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            value = row[col_idx].value
            if value:
                if value in values:
                    values[value].append(row_num)
                else:
                    values[value] = [row_num]
        
        for value, rows in values.items():
            if len(rows) > 1:
                duplicates.append({
                    'value': value,
                    'rows': rows,
                    'count': len(rows)
                })
        
        return duplicates
    
    def export_to_json(self, output_file: str, sheet_name: str = None):
        data = self.extract_data(sheet_name)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"Data exported to {output_file}")

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel Automation Tool')
    parser.add_argument('input', help='Input Excel file or directory')
    parser.add_argument('--output', help='Output file/directory')
    parser.add_argument('--action', choices=['extract', 'report', 'batch', 'duplicates', 'json'], default='extract')
    parser.add_argument('--sheet', help='Sheet name')
    parser.add_argument('--summary', nargs='+', help='Columns for summary')
    
    args = parser.parse_args()
    
    if args.action == 'batch':
        automator = ExcelAutomator("")
        count = automator.batch_process_files(args.input, args.output or "output")
        print(f"Processed {count} records from all files")
    
    elif os.path.isfile(args.input):
        automator = ExcelAutomator(args.input)
        
        if args.action == 'extract':
            data = automator.extract_data(args.sheet)
            print(f"Extracted {len(data)} records")
        
        elif args.action == 'report':
            data = automator.extract_data(args.sheet)
            automator.generate_report(data, args.output or "report.xlsx", args.summary)
        
        elif args.action == 'duplicates':
            if not args.summary:
                print("Please specify column name with --summary")
                return
            duplicates = automator.find_duplicates(args.summary[0], args.sheet)
            print(f"Found {len(duplicates)} duplicate groups")
            for d in duplicates[:5]:
                print(f"  Value: {d['value']}, Rows: {d['rows']}")
        
        elif args.action == 'json':
            automator.export_to_json(args.output or "output.json", args.sheet)
    
    else:
        print("Invalid input or action")

if __name__ == '__main__':
    main()